import pandas as pd
import numpy as np
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import seaborn as sns
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
import io
import os
import warnings
warnings.filterwarnings('ignore')

OUTPUT_PATH = r"D:\Yoobee College\DA\Activities\MSE-803-Data-Analytics\week 3\Data_Cleaning_Visualization_Mukesh.xlsx"
CSV_PATH    = r"C:\Users\Asus\Downloads\messy_dataset_Mukesh.csv"

# ── colours ──────────────────────────────────────────────────────────────────
HDR_FILL   = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
HDR_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name="Arial", bold=True, size=14, color="1F4E79")
SUB_FONT   = Font(name="Arial", bold=True, size=11, color="2E75B6")
BODY_FONT  = Font(name="Arial", size=10)
WARN_FILL  = PatternFill("solid", start_color="FFE699", end_color="FFE699")
ERR_FILL   = PatternFill("solid", start_color="FF9999", end_color="FF9999")
OK_FILL    = PatternFill("solid", start_color="C6EFCE", end_color="C6EFCE")
ALT_FILL   = PatternFill("solid", start_color="DEEAF1", end_color="DEEAF1")
thin       = Side(style="thin", color="BFBFBF")
BORDER     = Border(left=thin, right=thin, top=thin, bottom=thin)

def apply_header(ws, row, cols, texts, fill=HDR_FILL, font=HDR_FONT):
    for col, text in zip(cols, texts):
        c = ws.cell(row=row, column=col, value=text)
        c.fill, c.font, c.alignment, c.border = fill, font, Alignment(horizontal="center", vertical="center", wrap_text=True), BORDER

def style_cell(ws, row, col, value=None, fill=None, font=None, align=None, border=True, num_fmt=None):
    c = ws.cell(row=row, column=col)
    if value is not None:
        c.value = value
    if fill:  c.fill   = fill
    if font:  c.font   = font
    if align: c.alignment = align
    if border: c.border = BORDER
    if num_fmt: c.number_format = num_fmt
    return c

def set_col_widths(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

# ─────────────────────────────────────────────────────────────────────────────
# 1. RAW DATA
# ─────────────────────────────────────────────────────────────────────────────
raw_df = pd.read_csv(CSV_PATH, dtype=str, keep_default_na=False)

# Print raw data to console

print("RAW DATA")

print(raw_df.to_string(index=False))


# ─────────────────────────────────────────────────────────────────────────────
# 2. DOCUMENT ISSUES
# ─────────────────────────────────────────────────────────────────────────────
issues = [
    (2, "Bob",      "ID",        "2",               "Duplicate ID (same as row 3 Bob)",   "Removed duplicate row"),
    (3, "Bob",      "Salary",    "",                "Missing value",                       "Removed (duplicate of row 2)"),
    (4, "Charlie",  "Join Date", "",                "Missing value",                       "Marked as 'Unknown'"),
    (5, "David",    "Age",       "thirty-eight",    "Non-numeric text value",              "Converted to 38"),
    (6, "Eve",      "ID",        "",                "Missing value",                       "Assigned next ID: 6"),
    (6, "Eve",      "Join Date", "2019-13-01",      "Invalid date (month=13)",             "Marked as 'Unknown'"),
    (7, "Unknown",  "Name",      "",                "Missing value",                       "Filled as 'Unknown'"),
    (7, "Unknown",  "Salary",    "sixty five thousand", "Non-numeric text value",          "Converted to 65000"),
    (8, "Grace",    "Country",   "",                "Missing value",                       "Marked as 'Unknown'"),
    (9, "Heidi",    "Age",       "",                "Missing value",                       "Filled with median (30)"),
    (9, "Heidi",    "Salary",    "",                "Missing value",                       "Filled with median (61500)"),
]

# ─────────────────────────────────────────────────────────────────────────────
# 3. CLEAN DATA
# ─────────────────────────────────────────────────────────────────────────────
df = raw_df.copy()

# Fix Age column
df['Age'] = df['Age'].replace('thirty-eight', '38')
df['Age'] = pd.to_numeric(df['Age'], errors='coerce')

# Fix Salary column
df['Salary'] = df['Salary'].replace('sixty five thousand', '65000')
df['Salary'] = pd.to_numeric(df['Salary'], errors='coerce')

# Fix ID column
df['ID'] = pd.to_numeric(df['ID'], errors='coerce')

# Remove duplicate: keep first occurrence of ID=2 (row with age filled)
# Row index 1 has age missing; row index 2 has salary missing — merge them
# Row 0: Alice, Row 1: Bob(no age), Row 2: Bob(no salary) → keep one merged Bob
df.loc[1, 'Age'] = 30   # from row 2
df = df.drop(index=2).reset_index(drop=True)

# Assign missing ID for Eve
df.loc[df['Name'] == 'Eve', 'ID'] = 6

# Fill missing Name
df.loc[df['Name'] == '', 'Name'] = 'Unknown'

# Fix invalid / missing dates
df['Join Date'] = df['Join Date'].replace('', 'Unknown')
df.loc[df['Join Date'] == '2019-13-01', 'Join Date'] = 'Unknown'

# Fill missing Country
df['Country'] = df['Country'].replace('', 'Unknown')

# Compute medians on non-null numeric values for imputation
age_median    = int(df['Age'].dropna().median())
salary_median = int(df['Salary'].dropna().median())

# Fill missing Age & Salary with medians
df['Age']    = df['Age'].fillna(age_median)
df['Salary'] = df['Salary'].fillna(salary_median)

# Ensure ID is int
df['ID'] = df['ID'].astype(int)
df['Age'] = df['Age'].astype(int)
df['Salary'] = df['Salary'].astype(int)

clean_df = df.copy()

# ─────────────────────────────────────────────────────────────────────────────
# 4. STATISTICS & OUTLIER DETECTION (IQR method)
# ─────────────────────────────────────────────────────────────────────────────
def iqr_outliers(series):
    Q1, Q3 = series.quantile(0.25), series.quantile(0.75)
    IQR    = Q3 - Q1
    lower, upper = Q1 - 1.5 * IQR, Q3 + 1.5 * IQR
    return lower, upper, series[(series < lower) | (series > upper)]

age_lo, age_hi, age_out    = iqr_outliers(clean_df['Age'])
sal_lo, sal_hi, sal_out    = iqr_outliers(clean_df['Salary'])

# ─────────────────────────────────────────────────────────────────────────────
# 5. BUILD CHARTS
# ─────────────────────────────────────────────────────────────────────────────
CHART_DIR = r"D:\Yoobee College\DA\Activities\MSE-803-Data-Analytics\week 3\charts"
os.makedirs(CHART_DIR, exist_ok=True)

palette = "#2E75B6"

# 5a. Pearson Correlation Heatmap (Age & Salary only)
numeric_cols = clean_df[['Age', 'Salary']]
corr_matrix  = numeric_cols.corr(method='pearson')

fig, ax = plt.subplots(figsize=(6, 5))
cmap = sns.diverging_palette(220, 10, as_cmap=True)
sns.heatmap(corr_matrix, annot=True, fmt=".3f", cmap=cmap, vmin=-1, vmax=1,
            center=0, square=True, linewidths=1, linecolor='white',
            annot_kws={"size": 16, "weight": "bold", "color": "white"}, ax=ax)
ax.set_title("Pearson Correlation Heatmap\n(Age & Salary)", fontsize=13,
             fontweight='bold', pad=12, color="#1F4E79")
ax.set_xticklabels(ax.get_xticklabels(), rotation=0, fontsize=12)
ax.set_yticklabels(ax.get_yticklabels(), rotation=0, fontsize=12)
plt.tight_layout()
heatmap_path = os.path.join(CHART_DIR, "heatmap.png")
fig.savefig(heatmap_path, dpi=150, bbox_inches='tight')
plt.close()

# 5b. Scatter plot: Age vs Salary with regression trend line
from scipy import stats as sp_stats

slope, intercept, r_value, p_value, _ = sp_stats.linregress(clean_df['Age'], clean_df['Salary'])
x_line = np.linspace(clean_df['Age'].min() - 1, clean_df['Age'].max() + 1, 100)
y_line = slope * x_line + intercept

fig, ax = plt.subplots(figsize=(9, 6))
ax.set_facecolor('#F8F9FA')
fig.patch.set_facecolor('white')

ax.scatter(clean_df['Age'], clean_df['Salary'],
           color='#2A9D8F', edgecolors='white', linewidths=0.8,
           s=120, zorder=3)

for _, row in clean_df.iterrows():
    ax.annotate(row['Name'],
                xy=(row['Age'], row['Salary']),
                xytext=(6, 6), textcoords='offset points',
                fontsize=9, color='#333333')

ax.plot(x_line, y_line, color='#E76F51', linewidth=2.2,
        label=f'Trend  (r = {r_value:.2f})')

ax.set_xlabel("Age", fontsize=12)
ax.set_ylabel("Salary (NZD)", fontsize=12)
ax.set_title("Age vs Salary — Scatter & Regression",
             fontsize=14, fontweight='bold', color="#1F4E79", pad=14)
ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f'{x:,.0f}'))
ax.legend(fontsize=10, frameon=True)
ax.grid(True, alpha=0.4, color='white', linewidth=1.2)
ax.spines[['top', 'right']].set_visible(False)
plt.tight_layout()
scatter_path = os.path.join(CHART_DIR, "scatter_age_salary.png")
fig.savefig(scatter_path, dpi=150, bbox_inches='tight')
plt.close()

# 5c. Box plots for outlier detection
fig, axes = plt.subplots(1, 2, figsize=(10, 5))
for ax, col, lo, hi, out_idx, color in [
        (axes[0], 'Age',    age_lo,   age_hi,   age_out.index,   "#2E75B6"),
        (axes[1], 'Salary', sal_lo,   sal_hi,   sal_out.index,   "#C55A11")]:
    bp = ax.boxplot(clean_df[col], patch_artist=True, widths=0.5,
                    medianprops=dict(color='red', linewidth=2))
    bp['boxes'][0].set_facecolor(color + '55')
    bp['boxes'][0].set_edgecolor(color)
    # Mark outliers with names
    for idx in out_idx:
        val  = clean_df.loc[idx, col]
        name = clean_df.loc[idx, 'Name']
        ax.annotate(f" {name}\n({val})", xy=(1, val),
                    fontsize=8, color='red',
                    arrowprops=dict(arrowstyle='->', color='red'),
                    xytext=(1.18, val))
    ax.axhline(lo, color='orange', linestyle='--', linewidth=1, label=f'Lower fence ({lo:.0f})')
    ax.axhline(hi, color='purple', linestyle='--', linewidth=1, label=f'Upper fence ({hi:.0f})')
    ax.set_title(f'{col} – Outlier Detection (IQR)', fontweight='bold', color="#1F4E79")
    ax.set_ylabel(col)
    ax.legend(fontsize=8)
    ax.set_xticks([])
fig.suptitle("Box Plot Analysis – Outlier Detection", fontsize=13,
             fontweight='bold', color="#1F4E79")
plt.tight_layout()
boxplot_path = os.path.join(CHART_DIR, "boxplots.png")
fig.savefig(boxplot_path, dpi=150, bbox_inches='tight')
plt.close()

# 5c. Salary by Country bar chart (exclude 'Unknown' country)
fig, ax = plt.subplots(figsize=(7, 4))
country_sal = (clean_df[clean_df['Country'] != 'Unknown']
               .groupby('Country')['Salary']
               .mean()
               .sort_values(ascending=False))
palette = ["#2E75B6", "#C55A11", "#70AD47", "#7030A0", "#ED7D31"]
bars = ax.bar(country_sal.index, country_sal.values,
              color=palette[:len(country_sal)],
              edgecolor='white', linewidth=1.2)
ax.bar_label(bars, fmt='$%.0f', padding=3, fontsize=9)
ax.set_title("Average Salary by Country", fontweight='bold', color="#1F4E79")
ax.set_ylabel("Average Salary ($)")
ax.set_xlabel("Country")
ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f'${x:,.0f}'))
ax.spines[['top', 'right']].set_visible(False)
plt.tight_layout()
bar_path = os.path.join(CHART_DIR, "salary_by_country.png")
fig.savefig(bar_path, dpi=150, bbox_inches='tight')
plt.close()

# 5e. Missing Value Map: Before vs After Cleaning
raw_missing   = raw_df.replace('', np.nan).isnull()
clean_missing = clean_df.copy()
clean_missing['Join Date'] = clean_missing['Join Date'].replace('Unknown', np.nan)
clean_missing = clean_missing.isnull()

fig, axes = plt.subplots(1, 2, figsize=(14, 5))
fig.suptitle("Missing Value Map: Before vs After Cleaning",
             fontsize=14, fontweight='bold', color="#1F4E79", y=1.01)

for ax, data, title in [
        (axes[0], raw_missing,   "Raw (Before Cleaning)"),
        (axes[1], clean_missing, "Cleaned (After)")]:
    sns.heatmap(data, ax=ax, cbar=False,
                cmap=["#DEEAF1", "#FF6B6B"],
                yticklabels=False,
                linewidths=0.8, linecolor='white')
    ax.set_title(title, fontweight='bold', color="#1F4E79", fontsize=12)
    ax.set_xticklabels(ax.get_xticklabels(), rotation=45, ha='right', fontsize=10)

plt.tight_layout()
missing_map_path = os.path.join(CHART_DIR, "missing_map.png")
fig.savefig(missing_map_path, dpi=150, bbox_inches='tight')
plt.close()
#print(f"Missing map saved: {missing_map_path}")

# 5d. Age distribution histogram
fig, ax = plt.subplots(figsize=(7, 4))
ax.hist(clean_df['Age'], bins=6, color="#2E75B6", edgecolor='white', linewidth=1.2)
ax.axvline(clean_df['Age'].mean(),   color='red',    linestyle='--', linewidth=1.5, label=f"Mean ({clean_df['Age'].mean():.1f})")
ax.axvline(clean_df['Age'].median(), color='orange', linestyle='--', linewidth=1.5, label=f"Median ({clean_df['Age'].median():.1f})")
ax.set_title("Age Distribution", fontweight='bold', color="#1F4E79")
ax.set_xlabel("Age")
ax.set_ylabel("Frequency")
ax.legend()
ax.spines[['top', 'right']].set_visible(False)
plt.tight_layout()
hist_path = os.path.join(CHART_DIR, "age_distribution.png")
fig.savefig(hist_path, dpi=150, bbox_inches='tight')
plt.close()

# ─────────────────────────────────────────────────────────────────────────────
# 6. BUILD WORKBOOK
# ─────────────────────────────────────────────────────────────────────────────
wb = Workbook()

# ── Sheet 1: Raw Data ─────────────────────────────────────────────────────────
ws1 = wb.active
ws1.title = "1_Raw Data"
ws1.row_dimensions[1].height = 30
ws1.merge_cells("A1:F1")
c = ws1.cell(1, 1, "Raw Dataset – messy_dataset_Mukesh.csv  (Original, Unmodified)")
c.font, c.fill, c.alignment = TITLE_FONT, PatternFill("solid", start_color="DEEAF1", end_color="DEEAF1"), Alignment(horizontal="center", vertical="center")

headers = list(raw_df.columns)
apply_header(ws1, 2, range(1, len(headers)+1), headers)

issue_cells = {
    (3, 3): "Missing Age",
    (4, 5): "Missing Salary",
    (5, 6): "Missing Join Date",
    (6, 3): "Non-numeric: 'thirty-eight'",
    (7, 1): "Missing ID",
    (7, 6): "Invalid date: month=13",
    (8, 3): "Missing Name",
    (8, 5): "Non-numeric text",
    (9, 4): "Missing Country",
    (10,3): "Missing Age",
    (10,5): "Missing Salary",
}
for r_idx, row in enumerate(raw_df.itertuples(index=False), start=3):
    fill = ALT_FILL if r_idx % 2 == 0 else None
    for c_idx, val in enumerate(row, start=1):
        cell = ws1.cell(row=r_idx, column=c_idx, value=val if val != '' else None)
        cell.font, cell.border = BODY_FONT, BORDER
        cell.alignment = Alignment(horizontal="center")
        key = (r_idx, c_idx)
        if key in issue_cells:
            cell.fill = ERR_FILL
            cell.comment = None
        elif fill:
            cell.fill = fill

set_col_widths(ws1, {'A':6,'B':12,'C':8,'D':10,'E':10,'F':14})

ws1.cell(r_idx+2, 1, "Red cells = data quality issues identified during audit").font = Font(name="Arial", italic=True, color="FF0000", size=9)

# ── Sheet 2: Issues Log ───────────────────────────────────────────────────────
ws2 = wb.create_sheet("2_Issues Log")
ws2.merge_cells("A1:G1")
c = ws2.cell(1, 1, "Data Quality Audit – Issues Identified & Actions Taken")
c.font = TITLE_FONT
c.fill = PatternFill("solid", start_color="DEEAF1", end_color="DEEAF1")
c.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 30

apply_header(ws2, 2, range(1, 8),
             ["#", "Row", "Record", "Column", "Raw Value", "Issue Description", "Action Taken"])

issue_type_fills = {
    "Missing value":    WARN_FILL,
    "Duplicate":        PatternFill("solid", start_color="FFD7D7", end_color="FFD7D7"),
    "Non-numeric":      ERR_FILL,
    "Invalid date":     PatternFill("solid", start_color="FCE4D6", end_color="FCE4D6"),
}

for i, (row_num, record, col, raw_val, issue, action) in enumerate(issues, start=1):
    r = i + 2
    fill = ALT_FILL if i % 2 == 0 else None
    for c_idx, val in enumerate([i, row_num, record, col, raw_val if raw_val else "(empty)", issue, action], 1):
        cell = ws2.cell(row=r, column=c_idx, value=val)
        cell.font, cell.border = BODY_FONT, BORDER
        cell.alignment = Alignment(horizontal="left" if c_idx > 2 else "center", wrap_text=True)
        if fill: cell.fill = fill
    # Colour-code issue type
    issue_cell = ws2.cell(row=r, column=6)
    if "Missing" in issue:     issue_cell.fill = WARN_FILL
    elif "Duplicate" in issue: issue_cell.fill = PatternFill("solid", start_color="FFD7D7", end_color="FFD7D7")
    elif "Non-numeric" in issue: issue_cell.fill = ERR_FILL
    elif "Invalid" in issue:   issue_cell.fill = PatternFill("solid", start_color="FCE4D6", end_color="FCE4D6")

ws2.row_dimensions[2].height = 25
for r in range(3, len(issues)+3):
    ws2.row_dimensions[r].height = 22

set_col_widths(ws2, {'A':5,'B':6,'C':10,'D':12,'E':22,'F':32,'G':35})

# Legend
leg_row = len(issues) + 4
ws2.cell(leg_row, 1, "Legend:").font = SUB_FONT
legends = [("Yellow", "FFE699", "Missing value"), ("Pink", "FFD7D7", "Duplicate"),
           ("Red", "FF9999", "Non-numeric"), ("Orange", "FCE4D6", "Invalid date")]
for i, (lbl, col_hex, desc) in enumerate(legends):
    c = ws2.cell(leg_row + i + 1, 2)
    c.fill  = PatternFill("solid", start_color=col_hex, end_color=col_hex)
    c.border = BORDER
    ws2.cell(leg_row+i+1, 3, desc).font = BODY_FONT

# ── Sheet 3: Cleaned Data ─────────────────────────────────────────────────────
ws3 = wb.create_sheet("3_Cleaned Data")
ws3.merge_cells("A1:F1")
c = ws3.cell(1, 1, "Cleaned Dataset – After Data Quality Fixes")
c.font, c.fill, c.alignment = TITLE_FONT, PatternFill("solid", start_color="E2EFDA", end_color="E2EFDA"), Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[1].height = 30

apply_header(ws3, 2, range(1, 7),
             ["ID", "Name", "Age", "Country", "Salary", "Join Date"],
             fill=PatternFill("solid", start_color="375623", end_color="375623"))

for r_idx, row in enumerate(clean_df.itertuples(index=False), start=3):
    fill = PatternFill("solid", start_color="EBF3E8", end_color="EBF3E8") if r_idx % 2 == 0 else None
    for c_idx, val in enumerate(row, start=1):
        cell = ws3.cell(row=r_idx, column=c_idx, value=val)
        cell.font, cell.border = BODY_FONT, BORDER
        cell.alignment = Alignment(horizontal="center")
        if fill: cell.fill = fill
        if val == 'Unknown':
            cell.fill = WARN_FILL

set_col_widths(ws3, {'A':6,'B':12,'C':8,'D':12,'E':12,'F':14})

# Summary stats below cleaned table
sr = len(clean_df) + 4
ws3.cell(sr,   1, "Summary Statistics").font = SUB_FONT
ws3.merge_cells(f"A{sr}:F{sr}")

apply_header(ws3, sr+1, range(1, 6), ["Metric", "Count", "Mean", "Min", "Max"],
             fill=PatternFill("solid", start_color="375623", end_color="375623"))

for i, col in enumerate(['Age', 'Salary']):
    r = sr + 2 + i
    stats = [col, len(clean_df), round(clean_df[col].mean(), 1),
             clean_df[col].min(), clean_df[col].max()]
    for c_idx, v in enumerate(stats, 1):
        cell = ws3.cell(row=r, column=c_idx, value=v)
        cell.font, cell.border = BODY_FONT, BORDER
        cell.alignment = Alignment(horizontal="center")
        if i % 2 == 0: cell.fill = PatternFill("solid", start_color="EBF3E8", end_color="EBF3E8")

# ── Sheet 4: Correlation & Outliers ──────────────────────────────────────────
ws4 = wb.create_sheet("4_Correlation & Outliers")
ws4.merge_cells("A1:H1")
c = ws4.cell(1, 1, "Pearson Correlation Analysis & Outlier Detection")
c.font, c.fill, c.alignment = TITLE_FONT, PatternFill("solid", start_color="DEEAF1", end_color="DEEAF1"), Alignment(horizontal="center", vertical="center")
ws4.row_dimensions[1].height = 30

# Correlation matrix table
ws4.cell(3, 1, "Pearson Correlation Matrix").font = SUB_FONT
apply_header(ws4, 4, range(1, 4), ["", "Age", "Salary"])

corr_cols = ['Age', 'Salary']
for r_i, row_lbl in enumerate(corr_cols):
    r = r_i + 5
    ws4.cell(r, 1, row_lbl).font = Font(name="Arial", bold=True, size=10)
    ws4.cell(r, 1).border = BORDER
    ws4.cell(r, 1).alignment = Alignment(horizontal="center")
    for c_i, col_lbl in enumerate(corr_cols):
        val  = round(corr_matrix.loc[row_lbl, col_lbl], 4)
        cell = ws4.cell(row=r, column=c_i+2, value=val)
        cell.border, cell.font = BORDER, BODY_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.number_format = "0.0000"
        # colour by strength
        if abs(val) == 1.0:
            cell.fill = PatternFill("solid", start_color="BDD7EE", end_color="BDD7EE")
        elif abs(val) >= 0.7:
            cell.fill = OK_FILL
        elif abs(val) >= 0.4:
            cell.fill = WARN_FILL
        else:
            cell.fill = PatternFill("solid", start_color="F4CCCC", end_color="F4CCCC")

# Interpretation
ws4.cell(9, 1, "Interpretation:").font = SUB_FONT
interp = [
    ("Age vs Salary", round(corr_matrix.loc['Age','Salary'], 4), "Moderate positive – older employees tend to earn more"),
]
apply_header(ws4, 10, range(1, 4), ["Feature Pair", "Pearson r", "Interpretation"])
for i, (pair, r_val, desc) in enumerate(interp):
    r = 11 + i
    fill = ALT_FILL if i % 2 == 0 else None
    for c_idx, val in enumerate([pair, r_val, desc], 1):
        cell = ws4.cell(row=r, column=c_idx, value=val)
        cell.font, cell.border = BODY_FONT, BORDER
        cell.alignment = Alignment(horizontal="left" if c_idx != 2 else "center", wrap_text=True)
        if fill: cell.fill = fill

# Outlier summary
ws4.cell(15, 1, "Outlier Detection (IQR Method)").font = SUB_FONT
apply_header(ws4, 16, range(1, 7),
             ["Column", "Q1", "Q3", "IQR", "Lower Fence", "Upper Fence"])
for i, (col, lo, hi) in enumerate([
        ('Age',    age_lo,  age_hi),
        ('Salary', sal_lo,  sal_hi)]):
    r = 17 + i
    col_data = clean_df[col]
    Q1, Q3   = col_data.quantile(0.25), col_data.quantile(0.75)
    IQR_val  = Q3 - Q1
    fill = ALT_FILL if i % 2 == 0 else None
    for c_idx, val in enumerate([col, Q1, Q3, IQR_val, lo, hi], 1):
        cell = ws4.cell(row=r, column=c_idx, value=val if isinstance(val, str) else round(float(val), 2))
        cell.font, cell.border = BODY_FONT, BORDER
        cell.alignment = Alignment(horizontal="center")
        if fill: cell.fill = fill

ws4.cell(20, 1, "Outlier Records:").font = SUB_FONT
all_out = pd.concat([
    clean_df.loc[age_out.index].assign(Flag="Age outlier"),
    clean_df.loc[sal_out.index].assign(Flag="Salary outlier")
]).drop_duplicates()

if len(all_out) == 0:
    ws4.cell(21, 1, "No outliers detected in the cleaned dataset.").font = Font(name="Arial", color="375623", bold=True)
else:
    apply_header(ws4, 21, range(1, 8), list(clean_df.columns) + ["Flag"])
    for i, row_data in enumerate(all_out.itertuples(index=False)):
        r = 22 + i
        for c_idx, val in enumerate(row_data, 1):
            cell = ws4.cell(row=r, column=c_idx, value=val)
            cell.font, cell.border, cell.fill = BODY_FONT, BORDER, ERR_FILL
            cell.alignment = Alignment(horizontal="center")

set_col_widths(ws4, {'A':16,'B':10,'C':10,'D':10,'E':14,'F':14,'G':14,'H':25})

# ── Sheet 5: Visualizations ───────────────────────────────────────────────────
ws5 = wb.create_sheet("5_Visualizations")
ws5.merge_cells("A1:R1")
c = ws5.cell(1, 1, "Data Visualizations – Correlation Heatmap, Outlier Analysis & Distribution")
c.font, c.fill, c.alignment = TITLE_FONT, PatternFill("solid", start_color="DEEAF1", end_color="DEEAF1"), Alignment(horizontal="center", vertical="center")
ws5.row_dimensions[1].height = 30

ws5.cell(2, 1, "Fig 1 – Pearson Correlation Heatmap").font = SUB_FONT
ws5.cell(2, 10, "Fig 2 – Age vs Salary Scatter & Regression").font = SUB_FONT

img1 = XLImage(heatmap_path)
img1.anchor = "A3"
img1.width, img1.height = 380, 310
ws5.add_image(img1)

img2 = XLImage(scatter_path)
img2.anchor = "J3"
img2.width, img2.height = 560, 370
ws5.add_image(img2)

ws5.cell(24, 1, "Fig 3 – Box Plots (IQR Outlier Detection)").font = SUB_FONT
ws5.cell(24, 10, "Fig 4 – Average Salary by Country").font = SUB_FONT

img3 = XLImage(boxplot_path)
img3.anchor = "A25"
img3.width, img3.height = 540, 300
ws5.add_image(img3)

img4 = XLImage(bar_path)
img4.anchor = "J25"
img4.width, img4.height = 430, 300
ws5.add_image(img4)

ws5.cell(44, 1, "Fig 5 – Age Distribution").font = SUB_FONT
ws5.cell(44, 10, "Fig 6 – Missing Value Map: Before vs After Cleaning").font = SUB_FONT

img_hist = XLImage(hist_path)
img_hist.anchor = "A45"
img_hist.width, img_hist.height = 430, 300
ws5.add_image(img_hist)

img5 = XLImage(missing_map_path)
img5.anchor = "J45"
img5.width, img5.height = 560, 300
ws5.add_image(img5)

# ── Final save ────────────────────────────────────────────────────────────────
wb.save(OUTPUT_PATH)

print(f"\nCleaned dataset ({len(clean_df)} rows):")
print(clean_df.to_string(index=False))
print(f"\nAge median used for imputation : {age_median}")
print(f"Salary median used for imputation: {salary_median}")
print(f"\nPearson Correlation Matrix:")
print(corr_matrix.round(4))
print(f"\nOutliers – Age   : {list(age_out)}")
print(f"Outliers – Salary: {list(sal_out)}")
